from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys
# from bs4 import BeautifulSoup
# import requests
import time
import sys

from openpyxl.workbook import Workbook

# ExcelWriter,里面封装好了对Excel的写操作
from openpyxl.writer.excel import ExcelWriter

# get_column_letter函数将数字转换为相应的字母，如1-->A,2-->B

wb = Workbook()

# ew = ExcelWriter(workbook=wb)

# dest_file = "/Users/2fx0one/PycharmProjects/selenium/carInfo.xlsx"

ws = wb.active




# sys.exit(0)


# webdriver.Firefox



#数据采集
# list = []
# groupMap = {}
# allCarMapByGroup = {}
#
# for group in browser.find_elements_by_xpath("//div[@class='brands']/div"):
#     # for a in group.g
#     print(group.get_attribute("data-v-5d49e36e"))
#     groupName = group.find_element_by_class_name("floor")
#     print("组名：" + groupName.text)
#
#     # print(group.text)
#     for brand in group.find_elements_by_class_name("brand"):
#         print("品牌 " + brand.text)
#
#         list.append(brand.text)
#         groupMap[brand.text] = groupName.text
#
#         if allCarMapByGroup.get(groupName.text) is None:
#             allCarMapByGroup[groupName.text] = []
#
#         allCarMapByGroup[groupName.text].append(brand.text)
#
#
#     # break



list = ['Alfa·Romeo/阿尔法·罗密欧', '奥拓', '奥迪', '安驰', 'BORGWARD/宝沃', '保时捷', '别克', '北京', '北汽', '奔腾', '奔驰', '宝马', '宝骏', '宝龙',
        '宾利', '巴博斯', '布加迪', '本田', '标致', '比亚迪', '传祺', '昌河', '长城', '长安', '长安商用', 'DS', '东南', '东风', '东风小康', '东风风度',
        '大众', '大发', '大宇', '大迪', '大通', '道奇', '丰田', '富奇', '法拉利', '福特', '福田', '福迪', '菲亚特', '风神', '风行', 'GMC', '光冈',
        '观致', '华北', '华普', '华泰', '华阳', '华颂', '哈弗', '哈飞', '幻速', '恒天', '悍马', '汇众', '汉腾', '海格', '海马', '红旗', '黄海', '黑豹',
        'JEEP', '九龙', '佳星', '吉利', '吉奥', '捷豹', '江南', '江淮', '江铃', '解放', '金杯', '金程', '克莱斯勒', '凯翼', '凯迪拉克', '卡威', '卡尔森', '开瑞',
        '科尼赛克', '兰博基尼', '力帆', '劳伦士', '劳斯莱斯', '林肯', '猎豹', '理念', '莲花', '路特斯', '路虎', '铃木', '陆风', '雷克萨斯', '雷诺', '领克', 'MG',
        'mini', '玛莎拉蒂', '美亚', '迈凯伦', '迈巴赫', '马自达', '纳智捷', '欧宝', '欧朗', '讴歌', 'PAJNI/帕加尼', '启腾', '启辰', '奇瑞', '庆铃', '起亚', 'RUF',
        '日产', '瑞麒', '荣威', 'SWM/斯威', 'Smart', 'Springo', '三菱', '世爵', '双环', '双龙', '思铭', '斯巴鲁', '斯柯达', '绅宝', '萨博', '赛宝', '陕汽通家',
        'Tesla/特斯拉', '天马', '腾势', '通用', '通田', 'WEY（汽车）', '万丰', '五十铃', '五菱', '威兹曼', '威旺', '威麟', '沃尔沃', '新凯', '新雅途', '现代',
        '西雅特', '雪佛兰', '雪铁龙', '一汽', '云度', '依维柯', '扬子', '永源', '英致', '英菲尼迪', '蔚来(汽车)', '野马', 'ZD/知豆', '中兴', '中华', '中欧', '中顺', '众泰']
groupMap = {'Alfa·Romeo/阿尔法·罗密欧': 'A', '奥拓': 'A', '奥迪': 'A', '安驰': 'A', 'BORGWARD/宝沃': 'B', '保时捷': 'B', '别克': 'B', '北京': 'B', '北汽': 'B', '奔腾': 'B',
            '奔驰': 'B', '宝马': 'B', '宝骏': 'B', '宝龙': 'B', '宾利': 'B', '巴博斯': 'B', '布加迪': 'B', '本田': 'B', '标致': 'B', '比亚迪': 'B', '传祺': 'C', '昌河': 'C', '长城': 'C',
            '长安': 'C', '长安商用': 'C', 'DS': 'D', '东南': 'D', '东风': 'D', '东风小康': 'D', '东风风度': 'D', '大众': 'D', '大发': 'D', '大宇': 'D', '大迪': 'D', '大通': 'D',
            '道奇': 'D', '丰田': 'F', '富奇': 'F', '法拉利': 'F', '福特': 'F', '福田': 'F', '福迪': 'F', '菲亚特': 'F', '风神': 'F', '风行': 'F', 'GMC': 'G', '光冈': 'G', '观致':
                'G', '华北': 'H', '华普': 'H', '华泰': 'H', '华阳': 'H', '华颂': 'H', '哈弗': 'H', '哈飞': 'H', '幻速': 'H', '恒天': 'H', '悍马': 'H', '汇众': 'H', '汉腾': 'H', '海格': 'H',
            '海马': 'H', '红旗': 'H', '黄海': 'H', '黑豹': 'H', 'JEEP': 'J', '九龙': 'J', '佳星': 'J', '吉利': 'J', '吉奥': 'J', '捷豹': 'J', '江南': 'J', '江淮': 'J', '江铃': 'J',
            '解放': 'J', '金杯': 'J', '金程': 'J', '克莱斯勒': 'K', '凯翼': 'K', '凯迪拉克': 'K', '卡威': 'K', '卡尔森': 'K', '开瑞': 'K', '科尼赛克': 'K', '兰博基尼': 'L', '力帆': 'L',
            '劳伦士': 'L', '劳斯莱斯': 'L', '林肯': 'L', '猎豹': 'L', '理念': 'L', '莲花': 'L', '路特斯': 'L', '路虎': 'L', '铃木': 'L', '陆风': 'L', '雷克萨斯': 'L', '雷诺': 'L', '领克': 'L',
            'MG': 'M', 'mini': 'M', '玛莎拉蒂': 'M', '美亚': 'M', '迈凯伦': 'M', '迈巴赫': 'M', '马自达': 'M', '纳智捷': 'N', '欧宝': 'O', '欧朗': 'O', '讴歌': 'O', 'PAJNI/帕加尼': 'P', '启腾': 'Q',
            '启辰': 'Q', '奇瑞': 'Q', '庆铃': 'Q', '起亚': 'Q', 'RUF': 'R', '日产': 'R', '瑞麒': 'R', '荣威': 'R', 'SWM/斯威': 'S', 'Smart': 'S', 'Springo': 'S', '三菱': 'S', '世爵': 'S', '双环': 'S',
            '双龙': 'S', '思铭': 'S', '斯巴鲁': 'S', '斯柯达': 'S', '绅宝': 'S', '萨博': 'S', '赛宝': 'S', '陕汽通家': 'S', 'Tesla/特斯拉': 'T', '天马': 'T', '腾势': 'T', '通用': 'T', '通田': 'T', 'WEY（汽车）': 'W',
            '万丰': 'W', '五十铃': 'W', '五菱': 'W', '威兹曼': 'W', '威旺': 'W', '威麟': 'W', '沃尔沃': 'W', '新凯': 'X', '新雅途': 'X', '现代': 'X', '西雅特': 'X', '雪佛兰': 'X', '雪铁龙': 'X', '一汽': 'Y', '云度': 'Y',
            '依维柯': 'Y', '扬子': 'Y', '永源': 'Y', '英致': 'Y', '英菲尼迪': 'Y', '蔚来(汽车)': 'Y', '野马': 'Y', 'ZD/知豆': 'Z', '中兴': 'Z', '中华': 'Z', '中欧': 'Z', '中顺': 'Z', '众泰': 'Z'}

allCarMapByGroup = {
    'A': ['Alfa·Romeo/阿尔法·罗密欧', '奥拓', '奥迪', '安驰'],
    'B': ['BORGWARD/宝沃', '保时捷', '别克', '北京', '北汽', '奔腾', '奔驰', '宝马', '宝骏', '宝龙', '宾利', '巴博斯', '布加迪', '本田', '标致', '比亚迪'],
    'C': ['传祺', '昌河', '长城', '长安', '长安商用'],
    'D': ['DS', '东南', '东风', '东风小康', '东风风度', '大众', '大发', '大宇', '大迪', '大通', '道奇'],
    'F': ['丰田', '富奇', '法拉利', '福特', '福田', '福迪', '菲亚特', '风神', '风行'],
    'G': ['GMC', '光冈', '观致'],
    'H': ['华北', '华普', '华泰', '华阳', '华颂', '哈弗', '哈飞', '幻速', '恒天', '悍马', '汇众', '汉腾', '海格', '海马', '红旗', '黄海', '黑豹'],
    'J': ['JEEP', '九龙', '佳星', '吉利', '吉奥', '捷豹', '江南', '江淮', '江铃', '解放', '金杯', '金程'],
    'K': ['克莱斯勒', '凯翼', '凯迪拉克', '卡威', '卡尔森', '开瑞', '科尼赛克'],
    'L': ['兰博基尼', '力帆', '劳伦士', '劳斯莱斯', '林肯', '猎豹', '理念', '莲花', '路特斯', '路虎', '铃木', '陆风', '雷克萨斯', '雷诺', '领克'],
    'M': ['MG', 'mini', '玛莎拉蒂', '美亚', '迈凯伦', '迈巴赫', '马自达'],
    'N': ['纳智捷'],
    'O': ['欧宝', '欧朗', '讴歌'],
    'P': ['PAJNI/帕加尼'],
    'Q': ['启腾', '启辰', '奇瑞', '庆铃', '起亚'],
    'R': ['RUF', '日产', '瑞麒', '荣威'],
    'S': ['SWM/斯威', 'Smart', 'Springo', '三菱', '世爵', '双环', '双龙', '思铭', '斯巴鲁', '斯柯达', '绅宝', '萨博', '赛宝', '陕汽通家'],
    'T': ['Tesla/特斯拉', '天马', '腾势', '通用', '通田'],
    'W': ['WEY（汽车）', '万丰', '五十铃', '五菱', '威兹曼', '威旺', '威麟', '沃尔沃'],
    'X': ['新凯', '新雅途', '现代', '西雅特', '雪佛兰', '雪铁龙'],
    'Y': ['一汽', '云度', '依维柯', '扬子', '永源', '英致', '英菲尼迪', '蔚来(汽车)', '野马'],
    'Z': ['ZD/知豆', '中兴', '中华', '中欧', '中顺', '众泰']
}
print(list)
print(groupMap)
print(allCarMapByGroup)







def exportToExcel(groupName, brand_list=[]):
    browser = webdriver.Chrome()

    browser.implicitly_wait(30)
    browser.maximize_window()
    browser.get(
        "https://pages.tmall.com/wow/car/act/car-record?1=1&redirectURL=https%3A%2F%2Fpages.tmall.com%2Fwow%2Fcar%2Fact%2Fmycar-home%3Fut_sk%3D1.WJGytu833REDAMPBvJYBIsz5_21380790_1531111371995.Copy.windvane%26scm%3D20140623.1.3.86%26sourceType%3Dother%26ttid%3D201200%2540taobao_iphone_7.10.0%26spm%3Da21rf.8164232.11.13%26suid%3D3912F5A8-6102-45B1-AAE7-D684BFE5A7D3%26un%3Daad7bdb7bb511bbc8bd7bab98ffeb263%26share_crt_v%3D1%26sp_tk%3D4oKsbnpxTzBBdTltTlbigqw%3D%26cpp%3D1%26shareurl%3Dtrue%26short_name%3Dh.3c86ovL%26app%3Dchrome&spm=a223c.9682137/N.record.0#/select")


    ws.title = "carInfo"
    row_index = 1
    title = ["Group", "brandName", "Series", "URL", "YearSeries", "CarType"]
    for i in range(len(title)):
        # x = chr(ord('A')+i) + str(row_index)
        # print(x)
        ws.cell(row_index, i+1, title[i])
        # ws[x] = title[i]

    wait = WebDriverWait(browser, 10)
    list = brand_list
    for brandName in list:
        sqlhead =groupMap.get(brandName) + " " + brandName + " "

        column_index = 1
        # x = chr(ord('A')+i) + str(row_index)
        # print(x)
        # ws[x] = title[i]

        # print(brandName)
        # ws.cell(row_index, column_index, groupMap.get(brandName))
        # column_index += 1
        #
        # ws.cell(row_index, column_index, brandName)
        # column_index += 1

        row_data_head = [groupMap.get(brandName), brandName]
        try:
            wait.until(EC.element_to_be_clickable((By.XPATH, '//div[contains(text(), "' + brandName + '")]')))
            brand = browser.find_element_by_xpath('//div[contains(text(), "' + brandName + '")]')
            print("进入子页面 " + brand.text)
            brand.click()
            time.sleep(0.3)
        finally:
            pass
        #图片

        #进入子页面
        series_list = [series.text for series in browser.find_elements_by_class_name("brand")]

        for seriesName in series_list:
            series = browser.find_element_by_xpath('//div[contains(text(), "' + seriesName + '")]')
            print(series.text)
            sql = sqlhead
            sql += series.text + " "
            ActionChains(browser).move_to_element(series).perform()
            time.sleep(0.3)

            series_parent = browser.find_element_by_xpath('//div[contains(text(), "' + seriesName + '")]/..')

            # series_parent.
            url = series_parent.find_element_by_tag_name("img").get_attribute("src")
            # print(url)
            sql += url + " "
            print("+++ " + sql + " +++++")
            row_data_2 = row_data_head + [seriesName, url]

            #进入年款 子页面
            print("进入年款 子页面 " + seriesName)
            series.click()
            time.sleep(0.3)
            year_series_list = [x.text for x in browser.find_elements_by_class_name("brand")]

            for year_series_name in year_series_list:
                year_series = browser.find_element_by_xpath('//div[contains(text(), "' + year_series_name + '")]')

                sql2 = sql + year_series.text + " "
                #进入车型 子页面
                print("进入车型 子页面 " + year_series.text)
                year_series.click()
                time.sleep(0.3)

                row_data_3 = row_data_2 + [year_series_name]
                # row_data_3.append(year_series_name)

                car_type_list = [x.text for x in browser.find_elements_by_class_name("brand")]

                for car_type_name in car_type_list:
                    sql3 = sql2 + car_type_name
                    print("======" + sql3 + "==========")
                    row_data_4 = row_data_3 + [car_type_name]
                    # row_data_4.append(car_type_name)
                    print(row_data_4)
                    row_index += 1
                    for i in range(len(row_data_4)):
                        ws.cell(row_index, i + 1, row_data_4[i])
                        # ws.add_image()



                #退出车型 返回父页面
                print("车型 返回父页面  " + year_series_name)
                parent = browser.find_element_by_xpath('//span[contains(text(), "' + year_series_name + '")]')
                parent.click()
                time.sleep(0.3)

                # p = browser.find_elements_by_tag_name("span")
                # print(len(p))
                #
                # print([x.text for x in p])
                #
                # parent = [x for x in p if x.text == year_series_name]
                #
                # print([x.text for x in parent])
                # # a = list(filter(lambda x: x.text == year_series_name, parent))
                # print(len(parent))
                # p[-2].click()
                # time.sleep(0.1)

            # 退出年款 返回父页面
            print("年款 返回父页面  " + seriesName)
            #存在 品牌和车系同名的情况
            if brandName == seriesName:
                parents = browser.find_elements_by_xpath('//span[contains(text(), "' + seriesName + '")]')
                print([x.text for x in parents])
                parents[-1].click()
            else:
                browser.find_element_by_xpath('//span[contains(text(), "' + seriesName + '")]').click()
            time.sleep(0.3)




                #返回年款 父页面
                # series.
                #年款 seriesYear for
                    #车型 carType for
                # if str(url).startswith("data:image"):
                    # detail.send_keys(Keys.PAGE_DOWN)
                    # browser.execute_script('window.scrollTo(0, document.body.scrollHeight)')
                    # action.send_keys(Keys.PAGE_DOWN).perform()
                    # time.sleep(1)
            # browser.find_element_by_xpath("//div[@class='info']/span").click()
            # time.sleep(1)


        # finally:
            # 返回到父页面
        parent = browser.find_element_by_xpath('//span[contains(text(), "' + brandName + '")]')
        print("返回到父页面 " + parent.text)
        parent.click()


        # time.sleep(1)


    wb.save("/Users/2fx0one/PycharmProjects/selenium/carInfo" + groupName + ".xlsx")
    browser.close()


# exportToExcel('A', ['Alfa·Romeo/阿尔法·罗密欧', '奥拓', '奥迪', '安驰'])
# exportToExcel('B', ['别克'])
if __name__ == "__main__":
    cmd, first_arg = sys.argv
    if first_arg is None:
        sys.exit(-1)
    exportToExcel(first_arg, allCarMapByGroup[first_arg])

# for group_name, brand_list in allCarMapByGroup.items():
#     print(group_name + ": " + str(brand_list))
#     exportToExcel(group_name, brand_list)





# for brand in browser.find_elements_by_class_name("brand"):
#     print("品牌 " + brand.text)
#     print("品牌URL " + brand.find_element_by_class_name("logoImg").get_attribute("src"))
#
#     #进入子页面
#     action = ActionChains(browser)
#     action.move_to_element(brand).click().perform()
#     time.sleep(2)
#
#     for detail in browser.find_elements_by_class_name("brand"):
#         print(detail.text)
#         print(detail.find_element_by_class_name("logoImg").get_attribute("src"))
#
#     # 回到父页面
#     action.move_to_element(browser.find_element_by_xpath("//div[@class='info']/span")).click().perform()
#     time.sleep(1)
#     index += 1
#     if index == 2 :
#         break

#
# #遍历品牌分组
# for group in browser.find_elements_by_xpath("//div[@class='brands']/div"):
#     # for a in group.g
#     print(group.get_attribute("data-v-5d49e36e"))
#     groupName = group.find_element_by_class_name("floor")
#     print("组名：" + groupName.text)
#
#     # print(group.text)
#     for brand in group.find_elements_by_class_name("brand"):
#         print("品牌 " + brand.text)
#         print("品牌URL " + brand.find_element_by_class_name("logoImg").get_attribute("src"))
#         action = ActionChains(browser)
#         action.move_to_element(brand).click().perform()
#         time.sleep(2)
#         action.move_to_element(browser.find_element_by_xpath("//div[@class='info']/span")).click().perform()


    # for brand in group.find_elements_by_xpath("//div[@class='brand']"):
    #     print("品牌名： " + brand.text)
    # print(group.text)

    # 每组中的元素
    # for brand in group.find_elements_by_xpath("//div[@class='floor']"):
    #     print(brand.text)

    # logoImg = e.find_element_by_xpath("//img[@class='logoImg']")
    # print(e.text)
    # print(e.size())
    # print(logoImg.get_attribute('src'))
    #
    # #发送点击事件
    # #在新页面获取class=brand的图片
    # break


# browser.close()

#
# soup = BeautifulSoup(browser.page_source, "lxml")
#
# print(soup.find_all(name="div", class_="brandName"))
#
# # print(browser.page_source)
#
# # print(browser.find_element_by_class_name("brandName"))

